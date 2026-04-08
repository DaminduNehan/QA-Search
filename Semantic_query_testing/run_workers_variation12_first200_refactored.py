"""
Batch semantic search against the Workers QA API for queries from
``Variation_12_B2C_Humanized_Examples.json``.

For each query, paginates until no products, until ``total_pages`` is reached,
or until ``MAX_PAGES_PER_QUERY``. Writes one Excel row per page to support
resume: re-run with the same ``OUT_XLSX_PATH`` to continue after interruption.

Environment overrides:
    LIMIT_QUERIES   — max JSON entries to process (default: ``LIMIT_QUERIES``).
    OUT_XLSX_PATH   — output workbook path (default: sibling ``.xlsx`` file).

Run: ``python run_workers_variation12_first200_refactored.py``
"""

from __future__ import annotations

import json
import os
import sys
import time
import urllib.parse
import urllib.request
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from precision_recall_dashboard import (
    METRIC_HEADERS,
    calculate_precision_recall_fields,
    ensure_results_headers,
    generate_precision_recall_dashboard,
)

# ---------------------------------------------------------------------------
# Paths and API defaults (sibling to this script unless env overrides output)
# ---------------------------------------------------------------------------

INPUT_JSON = Path(__file__).with_name("Variation_12_B2C_Humanized_Examples.json")
OUT_XLSX_DEFAULT = Path(__file__).with_name("variation_12_first200_workers_results.xlsx")

BASE_URL = "https://search-api-wurthbaer-qa.search-villvay.workers.dev/search"
PAGE_SIZE = 24
IS_FILTER_BY_BRAND = "false"
LIMIT_QUERIES = 200

# Throttle and safety: avoid hammering the API and unbounded pagination.
# Business rule: only fetch first two pages per query.
MAX_PAGES_PER_QUERY = 2
REQUEST_DELAY_S = 0.15
TIMEOUT_S = 30
RETRIES_PER_PAGE = 4
RETRY_BACKOFF_S = 0.6

# Column order must match existing workbooks for resume compatibility.
BASE_RESULT_HEADERS: list[str] = [
    "query",
    "page",
    "expected_results",
    "actual_results",
    "number_of_results",
    "total_results",
    "total_pages",
    "page_size",
]
RESULT_HEADERS: list[str] = BASE_RESULT_HEADERS + METRIC_HEADERS


def extract_product_titles(payload: dict) -> list[str]:
    """
    Flatten product display strings from API JSON.

    Handles products as strings or dicts with common title field names.
    """
    results = payload.get("results") if isinstance(payload, dict) else None
    products = results.get("products") if isinstance(results, dict) else None
    if not isinstance(products, list):
        return []

    out: list[str] = []
    for p in products:
        if isinstance(p, str):
            t = p.strip()
        elif isinstance(p, dict):
            t = (
                p.get("productTitle")
                or p.get("primaryProductTitle")
                or p.get("title")
                or p.get("name")
                or ""
            )
            t = str(t).strip()
        else:
            t = ""
        if t:
            out.append(t)
    return out


def fetch_page(query: str, page: int) -> dict:
    """
    GET one search results page; retry with linear backoff on failure.

    Returns a dict with raw ``payload``, normalized summary fields, extracted
    ``products`` (titles), and ``last_error`` (empty string on success).
    """
    params = {
        "page": str(page),
        "query": query,
        "isFilterByBrand": IS_FILTER_BY_BRAND,
        "pageSize": str(PAGE_SIZE),
    }
    url = f"{BASE_URL}?{urllib.parse.urlencode(params)}"
    req = urllib.request.Request(url, method="GET")
    req.add_header("accept", "application/json")
    req.add_header("content-type", "application/json")
    req.add_header(
        "user-agent",
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/145.0.0.0 Safari/537.36",
    )

    payload: dict | None = None
    last_error: str | None = None
    for attempt in range(RETRIES_PER_PAGE):
        try:
            with urllib.request.urlopen(req, timeout=TIMEOUT_S) as resp:
                raw = resp.read().decode("utf-8", errors="replace")
            payload = json.loads(raw)
            break
        except Exception as e:  # noqa: BLE001
            last_error = str(e)
            if attempt < RETRIES_PER_PAGE - 1:
                time.sleep(RETRY_BACKOFF_S * (attempt + 1))
                continue
            raise

    summary = payload.get("summary") if isinstance(payload, dict) else {}
    return {
        "payload": payload,
        "page": int(summary.get("page") or page),
        "pageSize": int(summary.get("pageSize") or PAGE_SIZE),
        "total": int(summary.get("total") or 0),
        "totalPages": int(summary.get("totalPages") or 0),
        "products": extract_product_titles(payload),
        "last_error": last_error or "",
    }


def _sheet_headers(ws: Worksheet) -> list[str]:
    """First row of the sheet as stripped strings."""
    return [str(c.value or "").strip() for c in ws[1]]


def _count_completed_queries(ws: Worksheet, headers: list[str]) -> tuple[int, str | None]:
    """
    How many distinct non-empty ``query`` values appear (resume progress).

    Returns (count, last_query_seen) for logging; start index is ``count``
    clamped to available queries.
    """
    try:
        q_col = headers.index("query") + 1
    except ValueError:
        return 0, None

    last_query: str | None = None
    seen: set[str] = set()
    for r in range(2, ws.max_row + 1):
        qv = ws.cell(r, q_col).value
        qs = str(qv or "").strip()
        if not qs:
            continue
        seen.add(qs)
        last_query = qs
    return len(seen), last_query


def expected_outputs_joined(item: dict) -> str:
    """Newline-joined expected strings from ``expected_outputs`` or ``base_query``."""
    expected_list = item.get("expected_outputs") or []
    if not expected_list:
        base_query = str(item.get("base_query") or "").strip()
        expected_list = [base_query] if base_query else []
    return "\n".join(str(x).strip() for x in expected_list if str(x).strip())


def row_values_for_headers(row_dict: dict[str, object], headers: list[str]) -> list[object]:
    """Order ``row_dict`` values to match ``headers``."""
    return [row_dict.get(h, "") for h in headers]


def append_result_row(
    ws: Worksheet,
    headers: list[str],
    *,
    query: str,
    page: int,
    expected_joined: str,
    products: list[str] | None,
    total_results: int,
    total_pages_value: int,
    page_size: int,
    error_message: str | None = None,
) -> None:
    """One Excel row: either successful page or error placeholder."""
    if error_message is not None:
        actual = f"ERROR: {error_message}"
        n_results = 0
    else:
        actual = "\n".join(products or [])
        n_results = len(products or [])

    row_dict: dict[str, object] = {
        "query": query,
        "page": page,
        "expected_results": expected_joined,
        "actual_results": actual,
        "number_of_results": n_results,
        "total_results": total_results,
        "total_pages": total_pages_value,
        "page_size": page_size,
    }
    row_dict.update(calculate_precision_recall_fields(expected_joined, products or []))
    ws.append(row_values_for_headers(row_dict, headers))


def configure_line_buffered_stdout() -> None:
    """So ``print(..., flush=True)`` behaves well when stdout is piped."""
    try:
        sys.stdout.reconfigure(line_buffering=True)
    except Exception:  # noqa: BLE001
        pass


def load_query_items(json_path: Path, limit: int) -> list[dict]:
    """Read JSON array and take the first ``limit`` objects."""
    data = json.loads(json_path.read_text(encoding="utf-8"))
    return data[:limit]


def open_workbook_for_run(
    out_xlsx: Path,
    *,
    total_queries: int,
) -> tuple[Workbook, Worksheet, int]:
    """
    Create a new results workbook or load existing one and compute resume index.

    Returns (workbook, worksheet, completed_query_count for resume).
    ``completed_query_count`` is the number of distinct queries already written;
    the slice start is ``min(completed_query_count, total_queries)``.
    """
    if out_xlsx.exists():
        wb = load_workbook(out_xlsx)
        ws = wb["results"] if "results" in wb.sheetnames else wb.active
        ensure_results_headers(ws, BASE_RESULT_HEADERS)
        wb.save(out_xlsx)
        completed_queries, _ = _count_completed_queries(ws, RESULT_HEADERS)
        start_at = min(completed_queries, total_queries)
        print(
            f"Resuming existing workbook: {out_xlsx} "
            f"(completed queries={completed_queries}, starting at query index {start_at + 1})",
            flush=True,
        )
        return wb, ws, completed_queries

    wb = Workbook()
    ws = wb.active
    ws.title = "results"
    ws.append(RESULT_HEADERS)
    wb.save(out_xlsx)
    print(f"Created workbook: {out_xlsx}", flush=True)
    return wb, ws, 0


def run_pagination_for_query(
    ws: Worksheet,
    headers: list[str],
    *,
    query: str,
    expected_joined: str,
) -> None:
    """
    Fetch all pages for one query and append rows until stop conditions.
    """
    page = 1
    total_pages: int | None = None

    while True:
        try:
            res = fetch_page(query, page)
            products = res["products"]
            if total_pages is None:
                total_pages = res["totalPages"] or 0

            print(
                f"  page {page}/{total_pages or '?'} -> {len(products)} results "
                f"(total={res['total']}, pageSize={res['pageSize']})",
                flush=True,
            )

            append_result_row(
                ws,
                headers,
                query=query,
                page=page,
                expected_joined=expected_joined,
                products=products,
                total_results=res["total"],
                total_pages_value=total_pages or res["totalPages"] or 0,
                page_size=res["pageSize"],
            )

            if not products:
                break
            if total_pages and page >= total_pages:
                break
            if page >= MAX_PAGES_PER_QUERY:
                print(
                    f"  stopping: reached MAX_PAGES_PER_QUERY={MAX_PAGES_PER_QUERY}",
                    flush=True,
                )
                break

            page += 1
            time.sleep(REQUEST_DELAY_S)

        except Exception as e:  # noqa: BLE001
            print(f"  ERROR page {page}: {e}", flush=True)
            append_result_row(
                ws,
                headers,
                query=query,
                page=page,
                expected_joined=expected_joined,
                products=None,
                total_results=0,
                total_pages_value=total_pages or 0,
                page_size=PAGE_SIZE,
                error_message=str(e),
            )
            break


def main() -> None:
    configure_line_buffered_stdout()

    if not INPUT_JSON.exists():
        raise SystemExit(f"Missing input file: {INPUT_JSON}")

    limit = int(os.environ.get("LIMIT_QUERIES", str(LIMIT_QUERIES)))
    queries = load_query_items(INPUT_JSON, limit)

    out_xlsx = Path(os.environ.get("OUT_XLSX_PATH", str(OUT_XLSX_DEFAULT)))

    wb, ws, completed_queries = open_workbook_for_run(
        out_xlsx, total_queries=len(queries)
    )
    start_idx = min(completed_queries, len(queries))

    for qi, item in enumerate(queries[start_idx:], start=start_idx + 1):
        query = str(item.get("query") or "").strip()
        expected_joined = expected_outputs_joined(item)

        print(f"\n=== Query {qi}/{len(queries)} ===", flush=True)
        print(query, flush=True)

        run_pagination_for_query(ws, RESULT_HEADERS, query=query, expected_joined=expected_joined)
        wb.save(out_xlsx)

    dashboard_path = generate_precision_recall_dashboard(out_xlsx)
    print(f"Saved dashboard: {dashboard_path}", flush=True)
    print(f"\nSaved Excel: {out_xlsx}", flush=True)


if __name__ == "__main__":
    main()
