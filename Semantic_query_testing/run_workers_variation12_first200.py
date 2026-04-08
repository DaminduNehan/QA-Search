import json
import os
import sys
import time
import urllib.parse
import urllib.request
from pathlib import Path

from openpyxl import Workbook, load_workbook
from precision_recall_dashboard import (
    METRIC_HEADERS,
    calculate_precision_recall_fields,
    ensure_results_headers,
    generate_precision_recall_dashboard,
)

INPUT_JSON = Path(__file__).with_name("Variation_12_B2C_Humanized_Examples.json")
OUT_XLSX_DEFAULT = Path(__file__).with_name("variation_12_first200_workers_results.xlsx")

BASE_URL = "https://search-api-wurthbaer-qa.search-villvay.workers.dev/search"
PAGE_SIZE = 24
IS_FILTER_BY_BRAND = "false"
LIMIT_QUERIES = 200

# Safety caps to prevent runaway output on very broad queries.
# Business rule: only fetch first two pages per query.
MAX_PAGES_PER_QUERY = 2
REQUEST_DELAY_S = 0.15
TIMEOUT_S = 30
RETRIES_PER_PAGE = 4
RETRY_BACKOFF_S = 0.6


def extract_product_titles(payload: dict) -> list[str]:
    results = payload.get("results") if isinstance(payload, dict) else None
    products = results.get("products") if isinstance(results, dict) else None
    if not isinstance(products, list):
        return []
    out: list[str] = []
    for p in products:
        if isinstance(p, str):
            t = p.strip()
        elif isinstance(p, dict):
            t = (
                p.get("productTitle")
                or p.get("primaryProductTitle")
                or p.get("title")
                or p.get("name")
                or ""
            )
            t = str(t).strip()
        else:
            t = ""
        if t:
            out.append(t)
    return out


def fetch_page(query: str, page: int) -> dict:
    params = {
        "page": str(page),
        "query": query,
        "isFilterByBrand": IS_FILTER_BY_BRAND,
        "pageSize": str(PAGE_SIZE),
    }
    url = f"{BASE_URL}?{urllib.parse.urlencode(params)}"
    req = urllib.request.Request(url, method="GET")
    req.add_header("accept", "application/json")
    req.add_header("content-type", "application/json")
    req.add_header(
        "user-agent",
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/145.0.0.0 Safari/537.36",
    )
    payload = None
    last_error: str | None = None
    for attempt in range(RETRIES_PER_PAGE):
        try:
            with urllib.request.urlopen(req, timeout=TIMEOUT_S) as resp:
                raw = resp.read().decode("utf-8", errors="replace")
            payload = json.loads(raw)
            break
        except Exception as e:  # noqa: BLE001
            last_error = str(e)
            if attempt < RETRIES_PER_PAGE - 1:
                time.sleep(RETRY_BACKOFF_S * (attempt + 1))
                continue
            raise

    summary = payload.get("summary") if isinstance(payload, dict) else {}
    return {
        "payload": payload,
        "page": int(summary.get("page") or page),
        "pageSize": int(summary.get("pageSize") or PAGE_SIZE),
        "total": int(summary.get("total") or 0),
        "totalPages": int(summary.get("totalPages") or 0),
        "products": extract_product_titles(payload),
        "last_error": last_error or "",
    }


def _sheet_headers(ws) -> list[str]:
    return [str(c.value or "").strip() for c in ws[1]]


def _resume_state(ws, headers: list[str]) -> tuple[int, str | None]:
    try:
        q_col = headers.index("query") + 1
    except ValueError:
        return 0, None
    last_query = None
    seen: set[str] = set()
    for r in range(2, ws.max_row + 1):
        qv = ws.cell(r, q_col).value
        qs = str(qv or "").strip()
        if not qs:
            continue
        seen.add(qs)
        last_query = qs
    return len(seen), last_query


def main() -> None:
    try:
        # Ensure live progress in terminals/pipes
        sys.stdout.reconfigure(line_buffering=True)
    except Exception:  # noqa: BLE001
        pass
    if not INPUT_JSON.exists():
        raise SystemExit(f"Missing input file: {INPUT_JSON}")

    data = json.loads(INPUT_JSON.read_text(encoding="utf-8"))
    limit = int(os.environ.get("LIMIT_QUERIES", str(LIMIT_QUERIES)))
    queries = data[:limit]

    out_xlsx = Path(
        os.environ.get("OUT_XLSX_PATH", str(OUT_XLSX_DEFAULT))
    )

    base_headers = [
        "query",
        "page",
        "expected_results",
        "actual_results",
        "number_of_results",
        "total_results",
        "total_pages",
        "page_size",
    ]
    headers = base_headers + METRIC_HEADERS

    start_idx = 0
    if out_xlsx.exists():
        wb = load_workbook(out_xlsx)
        ws = wb["results"] if "results" in wb.sheetnames else wb.active
        ensure_results_headers(ws, base_headers)
        wb.save(out_xlsx)
        completed_queries, _ = _resume_state(ws, headers)
        start_idx = min(completed_queries, len(queries))
        print(
            f"Resuming existing workbook: {out_xlsx} "
            f"(completed queries={completed_queries}, starting at query index {start_idx + 1})",
            flush=True,
        )
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "results"
        ws.append(headers)
        wb.save(out_xlsx)
        print(f"Created workbook: {out_xlsx}", flush=True)

    for qi, q in enumerate(queries[start_idx:], start=start_idx + 1):
        query = str(q.get("query") or "").strip()
        expected_list = q.get("expected_outputs") or []
        if not expected_list:
            base_query = str(q.get("base_query") or "").strip()
            expected_list = [base_query] if base_query else []
        expected_joined = "\n".join(str(x).strip() for x in expected_list if str(x).strip())

        print(f"\n=== Query {qi}/{len(queries)} ===", flush=True)
        print(query, flush=True)

        page = 1
        total_pages = None
        while True:
            try:
                res = fetch_page(query, page)
                products = res["products"]
                if total_pages is None:
                    total_pages = res["totalPages"] or 0
                print(
                    f"  page {page}/{total_pages or '?'} -> {len(products)} results "
                    f"(total={res['total']}, pageSize={res['pageSize']})",
                    flush=True,
                )

                row_dict = {
                    "query": query,
                    "page": page,
                    "expected_results": expected_joined,
                    "actual_results": "\n".join(products),
                    "number_of_results": len(products),
                    "total_results": res["total"],
                    "total_pages": total_pages or res["totalPages"] or 0,
                    "page_size": res["pageSize"],
                }
                row_dict.update(
                    calculate_precision_recall_fields(expected_joined, products)
                )
                ws.append([row_dict.get(h, "") for h in headers])

                # Stop conditions
                if not products:
                    break
                if total_pages and page >= total_pages:
                    break
                if page >= MAX_PAGES_PER_QUERY:
                    print(
                        f"  stopping: reached MAX_PAGES_PER_QUERY={MAX_PAGES_PER_QUERY}",
                        flush=True,
                    )
                    break

                page += 1
                time.sleep(REQUEST_DELAY_S)
            except Exception as e:  # noqa: BLE001
                print(f"  ERROR page {page}: {e}", flush=True)
                row_dict = {
                    "query": query,
                    "page": page,
                    "expected_results": expected_joined,
                    "actual_results": f"ERROR: {e}",
                    "number_of_results": 0,
                    "total_results": 0,
                    "total_pages": total_pages or 0,
                    "page_size": PAGE_SIZE,
                }
                row_dict.update(calculate_precision_recall_fields(expected_joined, []))
                ws.append([row_dict.get(h, "") for h in headers])
                break
        # Save after each query so progress persists.
        wb.save(out_xlsx)

    dashboard_path = generate_precision_recall_dashboard(out_xlsx)
    print(f"Saved dashboard: {dashboard_path}", flush=True)
    print(f"\nSaved Excel: {out_xlsx}", flush=True)


if __name__ == "__main__":
    main()
