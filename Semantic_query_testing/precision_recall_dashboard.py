from __future__ import annotations

import html
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

METRIC_HEADERS: list[str] = [
    "recall_issue_count",
    "recall_issue_product_list",
    "precision_count",
    "precision_product_list",
]


def _split_non_empty_lines(value: object) -> list[str]:
    text = str(value or "")
    return [line.strip() for line in text.splitlines() if line.strip()]


def _ordered_unique(items: list[str]) -> list[str]:
    seen: set[str] = set()
    out: list[str] = []
    for item in items:
        key = item.casefold()
        if key in seen:
            continue
        seen.add(key)
        out.append(item)
    return out


def calculate_precision_recall_fields(
    expected_joined: str, actual_products: list[str] | None
) -> dict[str, object]:
    """
    Precision count is intersection(expected, actual).
    Recall issue count is expected items missing from actual.
    """
    expected = _ordered_unique(_split_non_empty_lines(expected_joined))
    actual = _ordered_unique([str(x).strip() for x in (actual_products or []) if str(x).strip()])
    actual_keys = {x.casefold() for x in actual}

    precision_products = [x for x in expected if x.casefold() in actual_keys]
    recall_issue_products = [x for x in expected if x.casefold() not in actual_keys]

    return {
        "recall_issue_count": len(recall_issue_products),
        "recall_issue_product_list": "\n".join(recall_issue_products),
        "precision_count": len(precision_products),
        "precision_product_list": "\n".join(precision_products),
    }


def ensure_results_headers(ws: Worksheet, base_headers: list[str]) -> list[str]:
    """
    Upgrade old results sheet (without metric columns) in-place.
    """
    existing = [str(c.value or "").strip() for c in ws[1]]
    full_headers = base_headers + METRIC_HEADERS

    if existing == full_headers:
        return full_headers
    if existing == base_headers:
        for idx, name in enumerate(METRIC_HEADERS, start=len(base_headers) + 1):
            ws.cell(row=1, column=idx, value=name)
        return full_headers
    raise SystemExit("Existing workbook headers do not match expected results schema.")


def generate_precision_recall_dashboard(out_xlsx: Path) -> Path:
    """
    Build query-wise precision/recall issue dashboard artifact as HTML.
    """
    wb = load_workbook(out_xlsx)
    ws = wb["results"] if "results" in wb.sheetnames else wb.active
    headers = [str(c.value or "").strip() for c in ws[1]]

    idx = {name: i for i, name in enumerate(headers)}
    for required in ("query", "expected_results", "actual_results"):
        if required not in idx:
            raise SystemExit(f"Missing required column '{required}' in results sheet.")

    query_data: dict[str, dict[str, list[str]]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        query = str(row[idx["query"]] or "").strip()
        if not query:
            continue

        expected_items = _ordered_unique(_split_non_empty_lines(row[idx["expected_results"]]))
        actual_raw = _split_non_empty_lines(row[idx["actual_results"]])
        actual_items = _ordered_unique(
            [x for x in actual_raw if not x.startswith("ERROR:")]
        )

        bucket = query_data.setdefault(query, {"expected": [], "actual": []})
        bucket["expected"].extend(expected_items)
        bucket["actual"].extend(actual_items)

    rows_html: list[str] = []
    for query, data in query_data.items():
        expected = _ordered_unique(data["expected"])
        actual = _ordered_unique(data["actual"])
        actual_keys = {x.casefold() for x in actual}

        precision_products = [x for x in expected if x.casefold() in actual_keys]
        recall_issues = [x for x in expected if x.casefold() not in actual_keys]

        expected_count = len(expected)
        actual_count = len(actual)
        precision_count = len(precision_products)
        recall_issue_count = len(recall_issues)

        # Precision uses actual products as denominator.
        precision_pct = (precision_count / actual_count * 100.0) if actual_count else 0.0
        # Recall issue percentage uses expected products as denominator.
        recall_issue_pct = (
            (recall_issue_count / expected_count * 100.0) if expected_count else 0.0
        )

        rows_html.append(
            "<tr>"
            f"<td>{html.escape(query)}</td>"
            f"<td>{expected_count}</td>"
            f"<td>{actual_count}</td>"
            f"<td>{precision_count}</td>"
            f"<td>{recall_issue_count}</td>"
            f"<td>{precision_pct:.2f}%</td>"
            f"<td>{recall_issue_pct:.2f}%</td>"
            "</tr>"
        )

    dashboard_path = out_xlsx.with_name(f"{out_xlsx.stem}_precision_recall_dashboard.html")
    page = f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8" />
  <title>Precision and Recall Dashboard</title>
  <style>
    body {{ font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif; margin: 24px; }}
    h1 {{ margin: 0 0 8px 0; }}
    p {{ color: #555; margin-top: 0; }}
    table {{ border-collapse: collapse; width: 100%; font-size: 14px; }}
    th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; vertical-align: top; }}
    th {{ background: #f7f7f7; position: sticky; top: 0; }}
    tr:nth-child(even) {{ background: #fafafa; }}
  </style>
</head>
<body>
  <h1>Query-wise Precision and Recall Issue Dashboard</h1>
  <p>
    Precision count = intersection(expected, actual).<br/>
    Recall issue count = expected items missing in actual.
  </p>
  <table>
    <thead>
      <tr>
        <th>Query</th>
        <th>Expected Count</th>
        <th>Actual Unique Count</th>
        <th>Precision Count</th>
        <th>Recall Issue Count</th>
        <th>Precision %</th>
        <th>Recall Issue %</th>
      </tr>
    </thead>
    <tbody>
      {''.join(rows_html)}
    </tbody>
  </table>
</body>
</html>
"""
    dashboard_path.write_text(page, encoding="utf-8")
    return dashboard_path
